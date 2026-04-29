"""
MSU mDOM Optical FAT Run List — XLSX to JSON converter

Requires: openpyxl
Install:  pip install openpyxl

Usage:
    python xlsx_to_json.py <input.xlsx> <output.json>

Logic:
    - If a cell contains rich text with individual run numbers in RED font
      → selected_run = the red-font run (largest if multiple red runs exist)
    - If the entire cell is red (uniform font color)
      → selected_run = largest run number in the cell
    - If no red runs
      → selected_run = largest run number overall
    - If no run numbers at all
      → selected_run = null, notes = original cell text
"""

import re
import json
import sys
import glob

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

from pathlib import Path
home = str(Path.home())


def extract_nums(text: str) -> list[int]:
    """Return all integers found in a string."""
    return [int(n) for n in re.findall(r'\b(\d+)\b', str(text))]


def is_red(color_rgb: str | None) -> bool:
    """Return True if an RGB hex string (8 chars with alpha) represents red."""
    if not color_rgb or len(color_rgb) < 8:
        return False
    r = int(color_rgb[2:4], 16)
    g = int(color_rgb[4:6], 16)
    b = int(color_rgb[6:8], 16)
    return r >= 180 and g < 80 and b < 80


def get_font_color(cell) -> str | None:
    """Return RGB hex color of a plain cell's font, or None."""
    try:
        color = cell.font.color
        if color and color.type == 'rgb':
            return color.rgb
    except Exception:
        pass
    return None


def parse_run_cell(cell) -> tuple[list[int], list[int]]:
    """
    Parse a run cell and return (all_runs, red_runs).

    Handles two cases:
      1. Rich text  — individual text segments may have different font colors.
      2. Plain text — entire cell has a single font color.
    """
    val = cell.value
    red_runs: list[int] = []
    all_runs: list[int] = []

    if isinstance(val, CellRichText):
        # Rich text: inspect each segment separately
        for block in val:
            if isinstance(block, TextBlock):
                nums = extract_nums(str(block))
                all_runs.extend(nums)
                try:
                    color = block.font.color.rgb if (
                        block.font and block.font.color
                        and block.font.color.type == 'rgb'
                    ) else None
                except Exception:
                    color = None
                if color and is_red(color):
                    red_runs.extend(nums)

    elif isinstance(val, str):
        all_runs = extract_nums(val)
        # Uniform cell color — if the whole cell is red, all runs count as red
        cell_color = get_font_color(cell)
        if cell_color and is_red(cell_color):
            red_runs = list(all_runs)

    return all_runs, red_runs

def get_device_list(string,device,geometry_files):
    device_list = []
    for ifile in geometry_files:
        if string in ifile:
            with open(ifile, 'r') as f:
                data = json.load(f)
            for idev in data[0]["devices"]:
                if device in idev["production_id"]:
                    device_list.append(idev["production_id"])
    return device_list


upgrade_commissioning_scripts = home+"/research_ua/icecube/software/upgrade_commissioning_scripts/"

geometry_files = sorted(glob.glob(upgrade_commissioning_scripts+"/geometry/string_*geometry*.json"))



def xlsx_to_json(input_path: str, output_path: str) -> None:
    wb = load_workbook(input_path, rich_text=True)
    ws = wb.active

    result: dict = {}
    deployed_device_list = (get_device_list("88","mDOM",geometry_files) + 
    get_device_list("89","mDOM",geometry_files) + 
    get_device_list("90","mDOM",geometry_files) + 
    get_device_list("91","mDOM",geometry_files) + 
    get_device_list("92","mDOM",geometry_files))
    print(deployed_device_list)

    for row in ws.iter_rows(min_row=2):
        mdom_cell = row[0]
        run_cell  = row[1] if len(row) > 1 else None

        mdom = mdom_cell.value
        if not mdom or not isinstance(mdom, str):
            continue
        mdom = mdom.strip()
        # print(f"mDOM_{mdom}")
        if f"mDOM_{mdom}" not in deployed_device_list:
            continue

        if run_cell is None or run_cell.value is None:
            result[mdom] = {"selected_run": None}
            continue

        all_runs, red_runs = parse_run_cell(run_cell)

        # Determine selected run
        if red_runs:
            selected = max(red_runs)          # largest red-font run
        elif all_runs:
            selected = max(all_runs)          # fallback: largest run overall
        else:
            selected = None

        entry: dict = {"selected_run": selected}

        # Preserve original text as notes when no run numbers were found
        if selected is None and run_cell.value:
            entry["notes"] = str(run_cell.value).strip()

        result[mdom] = entry

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"✓ Wrote {len(result)} entries → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(len(sys.argv))
        print("Usage: python xlsx_to_json.py <input.xlsx> <output.json>")
        print("Usage: python mdom_fat_runlist_to_json.py ../data/MSU_mDOM_Optical_FAT_Run_List.xlsx MSU_mDOM_FAT_Run_List.json")
        sys.exit(1)

    xlsx_to_json(sys.argv[1], sys.argv[2])